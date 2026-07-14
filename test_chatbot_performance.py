"""
Script untuk Testing Performance Chatbot RAG
- Akurasi chatbot (Recall@K, MRR, Precision@K, NDCG)
- Insertion Time (waktu insert ke vector DB)
- Query Time (waktu respon chatbot)
"""

import time
import json
import os
import sys
import shutil
import gc
import platform
import psutil
from datetime import datetime
from typing import List, Dict, Tuple
import numpy as np
from tqdm import tqdm

# Tambahkan path untuk import
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from app.chatbot import RAGChatbot
from app.initialize import initialize_vectorstore, load_documents
from langchain_community.vectorstores import Chroma
from langchain_community.embeddings import HuggingFaceEmbeddings

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Installing required packages...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas", "openpyxl"])
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import psutil
    HAS_PSUTIL = True
except ImportError:
    HAS_PSUTIL = False
    print("⚠ psutil not found. Install with: pip install psutil")
    print("  RAM monitoring will be disabled, but testing will continue.\n")


# ==================== HELPER FUNCTIONS ====================

def get_system_info():
    """Get system information"""
    info = {
        "platform": platform.system(),
        "platform_release": platform.release(),
        "python_version": sys.version.split()[0],
    }
    
    if HAS_PSUTIL:
        try:
            info["total_ram_gb"] = psutil.virtual_memory().total / (1024**3)
            info["available_ram_gb"] = psutil.virtual_memory().available / (1024**3)
            info["used_ram_percent"] = psutil.virtual_memory().percent
            info["cpu_count"] = psutil.cpu_count()
        except Exception as e:
            pass
    
    return info


def check_ollama_health(timeout: int = 5) -> Dict:
    """Check Ollama health and model availability"""
    import requests
    
    health = {
        "ollama_running": False,
        "mistral_available": False,
        "error": None
    }
    
    try:
        # Check if Ollama is running
        response = requests.get("http://localhost:11434/api/tags", timeout=timeout)
        if response.status_code == 200:
            health["ollama_running"] = True
            
            # Check if mistral model is available
            tags_data = response.json()
            models = tags_data.get("models", [])
            for model in models:
                if "mistral" in model.get("name", "").lower():
                    health["mistral_available"] = True
                    break
            
            if not health["mistral_available"]:
                health["error"] = "Mistral model not loaded. Run: ollama pull mistral"
        else:
            health["error"] = f"Ollama API returned status {response.status_code}"
    
    except requests.exceptions.ConnectionError:
        health["error"] = "Cannot connect to Ollama. Make sure it's running: ollama serve"
    except requests.exceptions.Timeout:
        health["error"] = f"Ollama timeout (>{timeout}s). It may be loading a model or system is busy"
    except Exception as e:
        health["error"] = f"Error checking Ollama: {str(e)}"
    
    return health


def safe_rmtree(path: str, max_retries: int = 3, delay: float = 1.0) -> bool:
    """
    Safely remove directory tree with retry logic for Windows file locking
    """
    if not os.path.exists(path):
        return True
    
    for attempt in range(max_retries):
        try:
            # Force garbage collection to release file handles
            gc.collect()
            time.sleep(0.5)
            
            # Try to remove directory
            shutil.rmtree(path)
            return True
        except (OSError, PermissionError) as e:
            if attempt < max_retries - 1:
                print(f"  ⚠ Directory lock (attempt {attempt + 1}/{max_retries}), retrying in {delay}s...")
                time.sleep(delay)
            else:
                print(f"  ❌ Failed to remove directory after {max_retries} attempts: {str(e)}")
                return False
    
    return False


# ==================== CONFIGURATION ====================
PERSIST_DIRECTORY = "chroma_db"
TEST_RESULTS_DIR = "test_results"
EMBEDDING_MODEL = "sentence-transformers/paraphrase-multilingual-MiniLM-L12-v2"

# Test Parameters
NUM_TEST_QUESTIONS = 10
NUM_REPETITIONS = 10
TOP_K = 7


# ==================== TEST DATA ====================
# Ground truth untuk evaluasi - sesuaikan dengan data Anda
TEST_QUESTIONS_WITH_GROUND_TRUTH = [
    {
        "question": "Apa itu Pangandaran?",
        "ground_truth_keywords": ["Pangandaran", "wisata", "pantai", "destinasi"],
        "relevant_documents": ["wisata_lists.toon"]
    },
    {
        "question": "Berapa harga tiket masuk?",
        "ground_truth_keywords": ["harga", "tiket", "masuk", "biaya"],
        "relevant_documents": ["harga_tiket_lists.toon"]
    },
    {
        "question": "Apa saja hotel yang tersedia?",
        "ground_truth_keywords": ["hotel", "penginapan", "tempat menginap"],
        "relevant_documents": ["hotel_lists.toon"]
    },
    {
        "question": "Bagaimana cara transportasi ke Pangandaran?",
        "ground_truth_keywords": ["transportasi", "cara", "perjalanan", "kendaraan"],
        "relevant_documents": ["transportasi.txt"]
    },
    {
        "question": "Dimana lokasi kantor polisi dan puskesmas Pangandaran?",
        "ground_truth_keywords": ["polisi", "puskesmas", "kantor", "alamat", "jalan"],
        "relevant_documents": ["tempat_penting_lists.toon"]
    },
    {
        "question": "Berapa harga paket wisata?",
        "ground_truth_keywords": ["paket", "wisata", "harga", "tour", "jetski"],
        "relevant_documents": ["paket_wisata_lists.toon"]
    },
    {
        "question": "Layanan emergency contact apa saja yang tersedia?",
        "ground_truth_keywords": ["polisi", "puskesmas", "telkom", "bank", "telpon"],
        "relevant_documents": ["nomor_penting_lists.toon"]
    },
    {
        "question": "Objek wisata apa yang wajib dikunjungi?",
        "ground_truth_keywords": ["wisata", "objek", "kunjungi", "destinasi"],
        "relevant_documents": ["wisata_lists.toon"]
    },
    {
        "question": "Bagaimana kondisi jalan menuju Pangandaran?",
        "ground_truth_keywords": ["jalan", "kondisi", "akses", "transportasi"],
        "relevant_documents": ["transportasi.txt"]
    },
    {
        "question": "Berapa biaya transportasi dari Bandung ke Pangandaran?",
        "ground_truth_keywords": ["Bandung", "Pangandaran", "transportasi", "harga", "rupiah", "tiket"],
        "relevant_documents": ["transportasi.txt"]
    }
]


# ==================== METRICS FUNCTIONS ====================
class MetricsCalculator:
    """Kelas untuk menghitung metrik evaluasi IR"""
    
    @staticmethod
    def calculate_recall_at_k(relevant_keywords: List[str], retrieved_docs: List[str], k: int) -> float:
        """
        Recall@K = (keywords found in top-k docs) / (total keywords)
        Measures: what % of relevant keywords were found in retrieved documents
        """
        if len(relevant_keywords) == 0:
            return 0.0
        
        # Convert keywords to lowercase for matching
        keywords_lower = [kw.lower() for kw in relevant_keywords]
        
        # Check top-k documents
        top_k_docs = retrieved_docs[:k]
        
        # Find which keywords appear in any of the top-k documents
        found_keywords = set()
        for doc in top_k_docs:
            doc_lower = doc.lower()
            for keyword in keywords_lower:
                if keyword in doc_lower:
                    found_keywords.add(keyword)
        
        # Recall = found keywords / total keywords
        return len(found_keywords) / len(relevant_keywords)
    
    @staticmethod
    def calculate_precision_at_k(relevant_keywords: List[str], retrieved_docs: List[str], k: int) -> float:
        """
        Precision@K = (documents with relevant keywords) / k
        Measures: what % of top-k documents contain at least one relevant keyword
        """
        if k == 0:
            return 0.0
        
        keywords_lower = [kw.lower() for kw in relevant_keywords]
        top_k_docs = retrieved_docs[:k]
        
        # Count documents that contain at least one keyword
        relevant_docs = 0
        for doc in top_k_docs:
            doc_lower = doc.lower()
            for keyword in keywords_lower:
                if keyword in doc_lower:
                    relevant_docs += 1
                    break  # Count this doc only once
        
        return relevant_docs / k
    
    @staticmethod
    def calculate_mrr(relevant_keywords: List[str], retrieved_docs: List[str]) -> float:
        """
        MRR = 1 / (rank of first relevant document)
        Measures: at what position we find the first document with relevant keywords
        """
        keywords_lower = [kw.lower() for kw in relevant_keywords]
        
        # Find position of first document containing any keyword
        for idx, doc in enumerate(retrieved_docs):
            doc_lower = doc.lower()
            for keyword in keywords_lower:
                if keyword in doc_lower:
                    return 1.0 / (idx + 1)
        
        return 0.0
    
    @staticmethod
    def calculate_ndcg(relevant_scores: List[float], k: int = 10) -> float:
        """
        NDCG@K = DCG@K / IDCG@K
        relevant_scores: list of relevance scores (1.0 for relevant, 0.0 for non-relevant)
        """
        def dcg(scores, k):
            dcg_sum = 0.0
            for i, score in enumerate(scores[:k]):
                dcg_sum += score / np.log2(i + 2)  # log2(i+2) untuk position-based discount
            return dcg_sum
        
        # Calculate DCG
        actual_dcg = dcg(relevant_scores, k)
        
        # Calculate IDCG (ideal DCG)
        ideal_scores = sorted(relevant_scores, reverse=True)
        ideal_dcg = dcg(ideal_scores, k)
        
        if ideal_dcg == 0:
            return 0.0
        
        return actual_dcg / ideal_dcg


# ==================== TEST FUNCTIONS ====================
class ChatbotPerformanceTester:
    """Main class untuk melakukan performance testing"""
    
    def __init__(self):
        self.chatbot = None
        self.vectordb = None
        self.embedding = None
        self.metrics = MetricsCalculator()
        self.results = {
            "accuracy": [],
            "insertion_time": None,
            "query_times": []
        }
        
        # Buat direktori untuk hasil
        if not os.path.exists(TEST_RESULTS_DIR):
            os.makedirs(TEST_RESULTS_DIR)
    
    def test_insertion_time(self) -> float:
        """Test waktu insertion data ke vector DB"""
        print("\n" + "="*60)
        print("Testing Insertion Time")
        print("="*60)
        
        # Close existing connections
        self.vectordb = None
        self.chatbot = None
        gc.collect()
        time.sleep(0.5)
        
        # Backup existing db
        backup_dir = None
        if os.path.exists(PERSIST_DIRECTORY):
            print("Creating backup of existing vector DB...")
            backup_dir = f"{PERSIST_DIRECTORY}_backup_{int(time.time())}"
            try:
                shutil.copytree(PERSIST_DIRECTORY, backup_dir)
                print(f"✓ Backup created: {backup_dir}")
            except Exception as e:
                print(f"❌ Backup failed: {str(e)}")
                backup_dir = None
            
            # Remove original directory safely
            print("Removing original vector DB...")
            if not safe_rmtree(PERSIST_DIRECTORY):
                print("⚠ Warning: Could not fully remove original DB, continuing anyway...")
        
        try:
            # Measure insertion time
            print("Initializing and indexing vector DB...")
            start_time = time.time()
            initialize_vectorstore(force_rebuild=True)
            insertion_time = time.time() - start_time
            
            print(f"✓ Insertion time: {insertion_time:.2f} seconds")
            
            return insertion_time
        
        except Exception as e:
            print(f"❌ Error during insertion: {str(e)}")
            return 0.0
        
        finally:
            # Restore backup
            if backup_dir and os.path.exists(backup_dir):
                print("Restoring original vector DB from backup...")
                
                # Close any open connections
                self.vectordb = None
                self.chatbot = None
                gc.collect()
                time.sleep(0.5)
                
                # Remove the new db
                if os.path.exists(PERSIST_DIRECTORY):
                    safe_rmtree(PERSIST_DIRECTORY, max_retries=5, delay=1.0)
                
                # Restore backup
                try:
                    shutil.move(backup_dir, PERSIST_DIRECTORY)
                    print(f"✓ Restored from backup")
                except Exception as e:
                    print(f"❌ Failed to restore backup: {str(e)}")
    
    def test_query_time_and_accuracy(self):
        """Test query time dan akurasi chatbot"""
        print("\n" + "="*60)
        print("Testing Query Time & Accuracy")
        print("="*60)
        
        # Initialize chatbot
        print("Initializing chatbot...")
        self.chatbot = RAGChatbot()
        
        if not self.chatbot.ollama_available:
            print("⚠ Warning: Ollama not available. Accuracy test will be limited.")
        
        accuracy_results = []
        query_times = []
        
        # Seleksi test questions
        test_questions = TEST_QUESTIONS_WITH_GROUND_TRUTH[:NUM_TEST_QUESTIONS]
        
        for q_idx, test_data in enumerate(tqdm(test_questions, desc="Testing Questions")):
            question = test_data["question"]
            relevant_keywords = test_data["ground_truth_keywords"]
            
            print(f"\n[Q{q_idx+1}] {question}")
            
            q_times = []
            q_recalls = []
            q_precisions = []
            q_mrrs = []
            q_ndcgs = []
            
            for rep in range(NUM_REPETITIONS):
                try:
                    # Measure query time with timeout handling
                    start_time = time.time()
                    result = self.chatbot.ask(question)
                    query_time = time.time() - start_time
                    
                    q_times.append(query_time)
                    
                    # Extract retrieved documents
                    retrieved_docs = result.get("sources", [])
                    
                    # Check if answer contains error
                    answer = result.get("answer", "")
                    if "Ollama call failed" in answer or "Error:" in answer:
                        print(f"    ⚠ Rep {rep+1}/{NUM_REPETITIONS}: Ollama error - {answer[:50]}...")
                        # Still add default scores
                        q_recalls.append(0.0)
                        q_precisions.append(0.0)
                        q_mrrs.append(0.0)
                        q_ndcgs.append(0.0)
                        continue
                    
                    # Konversi ke list item untuk ranking
                    retrieved_items = retrieved_docs
                    
                    # Hitung relevansi scores (1.0 jika mengandung keyword, 0.0 jika tidak)
                    relevance_scores = []
                    for doc in retrieved_items:
                        score = 1.0 if any(keyword.lower() in doc.lower() for keyword in relevant_keywords) else 0.0
                        relevance_scores.append(score)
                    
                    # Hitung metrik dengan logic yang benar (keyword matching dalam documents)
                    recall = self.metrics.calculate_recall_at_k(
                        relevant_keywords,
                        retrieved_items,
                        TOP_K
                    )
                    precision = self.metrics.calculate_precision_at_k(
                        relevant_keywords,
                        retrieved_items,
                        TOP_K
                    )
                    mrr = self.metrics.calculate_mrr(
                        relevant_keywords,
                        retrieved_items
                    )
                    ndcg = self.metrics.calculate_ndcg(relevance_scores, k=TOP_K)
                    
                    q_recalls.append(recall)
                    q_precisions.append(precision)
                    q_mrrs.append(mrr)
                    q_ndcgs.append(ndcg)
                
                except Exception as e:
                    print(f"    ❌ Rep {rep+1}/{NUM_REPETITIONS}: Exception - {str(e)[:50]}...")
                    # Add default scores for failed query
                    q_times[-1:] = []  # Remove the added time if exception during processing
                    q_recalls.append(0.0)
                    q_precisions.append(0.0)
                    q_mrrs.append(0.0)
                    q_ndcgs.append(0.0)
            
            # Average per question (only if we have valid data)
            if q_times:
                avg_query_time = np.mean(q_times)
                avg_recall = np.mean(q_recalls) if q_recalls else 0.0
                avg_precision = np.mean(q_precisions) if q_precisions else 0.0
                avg_mrr = np.mean(q_mrrs) if q_mrrs else 0.0
                avg_ndcg = np.mean(q_ndcgs) if q_ndcgs else 0.0
            else:
                avg_query_time = 0.0
                avg_recall = 0.0
                avg_precision = 0.0
                avg_mrr = 0.0
                avg_ndcg = 0.0
            
            accuracy_results.append({
                "question": question,
                "avg_query_time": avg_query_time,
                "recall_at_k": avg_recall,
                "precision_at_k": avg_precision,
                "mrr": avg_mrr,
                "ndcg": avg_ndcg,
                "num_repetitions": NUM_REPETITIONS
            })
            
            query_times.extend(q_times)
            
            if avg_query_time > 0:
                print(f"  ├─ Avg Query Time: {avg_query_time*1000:.2f}ms")
                print(f"  ├─ Recall@{TOP_K}: {avg_recall:.4f}")
                print(f"  ├─ Precision@{TOP_K}: {avg_precision:.4f}")
                print(f"  ├─ MRR: {avg_mrr:.4f}")
                print(f"  └─ NDCG: {avg_ndcg:.4f}")
            else:
                print(f"  └─ ⚠ Query failed for all repetitions")
        
        self.results["accuracy"] = accuracy_results
        self.results["query_times"] = query_times
        
        return accuracy_results, query_times
    
    def generate_excel_report(self, insertion_time: float, accuracy_results: List[Dict], query_times: List[float]):
        """Generate Excel report"""
        print("\n" + "="*60)
        print("Generating Excel Report")
        print("="*60)
        
        # Create Excel workbook
        wb = Workbook()
        
        # ===== Sheet 1: Summary =====
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary section
        ws_summary["A1"] = "CHATBOT PERFORMANCE TEST REPORT"
        ws_summary["A1"].font = Font(bold=True, size=14)
        ws_summary.merge_cells("A1:D1")
        
        ws_summary["A2"] = f"Test Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws_summary.merge_cells("A2:D2")
        
        ws_summary["A4"] = "Test Configuration"
        ws_summary["A4"].font = Font(bold=True, size=11)
        
        ws_summary["A5"] = "Number of Questions:"
        ws_summary["B5"] = NUM_TEST_QUESTIONS
        ws_summary["A6"] = "Repetitions per Question:"
        ws_summary["B6"] = NUM_REPETITIONS
        ws_summary["A7"] = "Top-K for Metrics:"
        ws_summary["B7"] = TOP_K
        
        ws_summary["A9"] = "Overall Results"
        ws_summary["A9"].font = Font(bold=True, size=11)
        
        ws_summary["A10"] = "Insertion Time (seconds):"
        ws_summary["B10"] = insertion_time
        ws_summary["B10"].number_format = '0.00'
        
        avg_recall = np.mean([r["recall_at_k"] for r in accuracy_results]) if accuracy_results else 0
        avg_precision = np.mean([r["precision_at_k"] for r in accuracy_results]) if accuracy_results else 0
        avg_mrr = np.mean([r["mrr"] for r in accuracy_results]) if accuracy_results else 0
        avg_ndcg = np.mean([r["ndcg"] for r in accuracy_results]) if accuracy_results else 0
        avg_query_time = np.mean([r["avg_query_time"] for r in accuracy_results]) if accuracy_results else 0
        
        ws_summary["A11"] = "Avg Query Time (seconds):"
        ws_summary["B11"] = avg_query_time
        ws_summary["B11"].number_format = '0.0000'
        
        ws_summary["A12"] = "Avg Recall@K:"
        ws_summary["B12"] = avg_recall
        ws_summary["B12"].number_format = '0.0000'
        
        ws_summary["A13"] = "Avg Precision@K:"
        ws_summary["B13"] = avg_precision
        ws_summary["B13"].number_format = '0.0000'
        
        ws_summary["A14"] = "Avg MRR:"
        ws_summary["B14"] = avg_mrr
        ws_summary["B14"].number_format = '0.0000'
        
        ws_summary["A15"] = "Avg NDCG:"
        ws_summary["B15"] = avg_ndcg
        ws_summary["B15"].number_format = '0.0000'
        
        # Adjust column width
        ws_summary.column_dimensions["A"].width = 30
        ws_summary.column_dimensions["B"].width = 20
        
        # ===== Sheet 2: Detailed Results =====
        ws_detailed = wb.create_sheet("Detailed Results")
        
        # Headers
        headers = ["Question #", "Question", "Avg Query Time (s)", "Recall@K", "Precision@K", "MRR", "NDCG"]
        for col, header in enumerate(headers, 1):
            cell = ws_detailed.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Data rows
        for idx, result in enumerate(accuracy_results, 2):
            ws_detailed.cell(row=idx, column=1, value=idx-1)
            ws_detailed.cell(row=idx, column=2, value=result["question"])
            ws_detailed.cell(row=idx, column=3, value=result["avg_query_time"]).number_format = '0.0000'
            ws_detailed.cell(row=idx, column=4, value=result["recall_at_k"]).number_format = '0.0000'
            ws_detailed.cell(row=idx, column=5, value=result["precision_at_k"]).number_format = '0.0000'
            ws_detailed.cell(row=idx, column=6, value=result["mrr"]).number_format = '0.0000'
            ws_detailed.cell(row=idx, column=7, value=result["ndcg"]).number_format = '0.0000'
            
            for col in range(1, 8):
                ws_detailed.cell(row=idx, column=col).border = border
        
        # Adjust column width
        ws_detailed.column_dimensions["A"].width = 12
        ws_detailed.column_dimensions["B"].width = 40
        ws_detailed.column_dimensions["C"].width = 18
        ws_detailed.column_dimensions["D"].width = 14
        ws_detailed.column_dimensions["E"].width = 16
        ws_detailed.column_dimensions["F"].width = 12
        ws_detailed.column_dimensions["G"].width = 12
        
        # ===== Sheet 3: Query Time Statistics =====
        ws_stats = wb.create_sheet("Query Time Stats")
        
        if query_times:
            ws_stats["A1"] = "Query Time Statistics (milliseconds)"
            ws_stats["A1"].font = Font(bold=True, size=12)
            
            ws_stats["A3"] = "Metric"
            ws_stats["B3"] = "Value (ms)"
            ws_stats["A3"].fill = header_fill
            ws_stats["B3"].fill = header_fill
            ws_stats["A3"].font = header_font
            ws_stats["B3"].font = header_font
            
            query_times_ms = [t * 1000 for t in query_times]
            
            stats_data = [
                ("Minimum", np.min(query_times_ms)),
                ("Maximum", np.max(query_times_ms)),
                ("Mean", np.mean(query_times_ms)),
                ("Median", np.median(query_times_ms)),
                ("Std Dev", np.std(query_times_ms)),
                ("25th Percentile", np.percentile(query_times_ms, 25)),
                ("75th Percentile", np.percentile(query_times_ms, 75)),
                ("95th Percentile", np.percentile(query_times_ms, 95)),
            ]
            
            for idx, (metric, value) in enumerate(stats_data, 4):
                ws_stats.cell(row=idx, column=1, value=metric)
                ws_stats.cell(row=idx, column=2, value=value).number_format = '0.00'
                ws_stats.cell(row=idx, column=1).border = border
                ws_stats.cell(row=idx, column=2).border = border
            
            ws_stats.column_dimensions["A"].width = 20
            ws_stats.column_dimensions["B"].width = 20
        
        # Save workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{TEST_RESULTS_DIR}/chatbot_performance_report_{timestamp}.xlsx"
        wb.save(filename)
        
        print(f"✓ Report saved to: {filename}")
        return filename
    
    def run_all_tests(self):
        """Run semua test"""
        print("\n" + "="*60)
        print("CHATBOT PERFORMANCE TEST SUITE")
        print("="*60)
        print(f"Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        # Get system info
        sys_info = get_system_info()
        print(f"Python Version: {sys_info.get('python_version', 'N/A')}")
        print(f"OS: {sys_info.get('platform', 'N/A')} {sys_info.get('platform_release', 'N/A')}")
        
        if HAS_PSUTIL and 'total_ram_gb' in sys_info:
            print(f"System RAM: {sys_info['total_ram_gb']:.1f} GB")
            print(f"Available RAM: {sys_info['available_ram_gb']:.1f} GB ({100-sys_info['used_ram_percent']:.1f}% free)")
            print(f"CPU Cores: {sys_info.get('cpu_count', 'N/A')}")
            
            # Warn if RAM is low
            if sys_info['available_ram_gb'] < 4:
                print("\n⚠️  WARNING: Low available RAM (<4GB)")
                print("   Ollama may struggle to load models or crash during testing")
                print("   Consider closing other applications or using a model with lower memory requirements\n")
        
        # Check Ollama health
        print("\nChecking Ollama Health...")
        ollama_health = check_ollama_health()
        
        if ollama_health["ollama_running"]:
            print("✓ Ollama is running")
        else:
            print(f"❌ Ollama issue: {ollama_health['error']}")
            print("   Start Ollama with: ollama serve")
            return None
        
        if ollama_health["mistral_available"]:
            print("✓ Mistral model is available")
        else:
            print(f"⚠ Model issue: {ollama_health['error']}")
            print("   Pull model with: ollama pull mistral")
            return None
        
        try:
            # Test 1: Insertion Time
            print("\n[1/3] Running Insertion Time Test...")
            insertion_time = self.test_insertion_time()
            
            # Close connections after insertion test
            self.vectordb = None
            self.chatbot = None
            gc.collect()
            time.sleep(1)
            
            # Test 2: Query Time & Accuracy
            print("\n[2/3] Running Query Time & Accuracy Tests...")
            accuracy_results, query_times = self.test_query_time_and_accuracy()
            
            # Close connections after query test
            self.chatbot = None
            gc.collect()
            time.sleep(0.5)
            
            # Generate Report
            print("\n[3/3] Generating Excel Report...")
            report_file = self.generate_excel_report(insertion_time, accuracy_results, query_times)
            
            print("\n" + "="*60)
            print("✅ ALL TESTS COMPLETED SUCCESSFULLY")
            print("="*60)
            print(f"End Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"\n📊 Report saved to: {report_file}")
            
            return report_file
            
        except Exception as e:
            print(f"\n❌ Error during testing: {str(e)}")
            
            # Check if it's an Ollama-related error
            if "404" in str(e) or "Ollama" in str(e):
                print("\n⚠️  OLLAMA ERROR DETECTED:")
                print("   - Model may not be loaded into memory")
                print("   - Ollama process may have crashed due to low RAM")
                print("   - Try: ollama pull mistral")
                print("   - Monitor system RAM while Ollama is loading")
            
            import traceback
            traceback.print_exc()
            return None
        
        finally:
            # Cleanup
            self.vectordb = None
            self.chatbot = None
            gc.collect()


# ==================== MAIN ====================
if __name__ == "__main__":
    tester = ChatbotPerformanceTester()
    tester.run_all_tests()